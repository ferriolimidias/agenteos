from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, status, UploadFile, File, Form
import io
import csv
import json
import uuid
from datetime import datetime
import pdfplumber
import traceback
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import delete, update, func, or_
from typing import List
from fastapi.responses import Response

from db.database import get_db, AsyncSessionLocal
from sqlalchemy.orm import selectinload
from db.models import (
    ADMIN_EMPRESA_ROLE,
    ROOT_ADMIN_ROLE,
    CampanhaDisparo,
    CampanhaDisparoStatus,
    Empresa,
    Conhecimento,
    Usuario,
    ConhecimentoRAG,
    CRMFunil,
    CRMEtapa,
    CRMLead,
    MensagemHistorico,
    AgendaConfiguracao,
    AgendamentoLocal,
    Conexao,
    DestinosTransferencia,
    HistoricoTransferencia,
    TagGroup,
    TagCRM,
    TemplateMensagem,
    is_admin_empresa_role,
    is_root_admin_email,
    is_super_admin_role,
    normalize_user_email,
)
from pydantic import BaseModel, Field
from typing import Dict, Any, Optional

from app.core.security import get_password_hash
from app.services.campanha_service import extrair_variaveis_template, gerar_preview_campanha, processar_campanha_disparo
from app.services.ads_integration_service import notificar_conversao_ads
from app.services.tag_crm_service import (
    listar_tags_oficiais_ou_existentes,
    normalizar_tags,
)
from app.services.transferencia_service import executar_transferencia_atendimento, testar_destino_transferencia

from app.schemas import (
    EmpresaCreate,
    EmpresaResponse,
    EmpresaSetupRequest,
    EmpresaUpdate,
    EvolutionCredentials,
    IAConfigResponse,
    IAConfigUpdateRequest,
    StandardMessage,
)

router = APIRouter(
    prefix="/empresas",
    tags=["Empresas"]
)

SIMULADOR_LEAD_ID = "ID_TESTE_SIMULADOR"


def _parse_uuid_or_none(value: str | None) -> uuid.UUID | None:
    try:
        return uuid.UUID(str(value))
    except (ValueError, TypeError, AttributeError):
        return None


async def _limpar_estado_simulador_redis(sessao_id: str = SIMULADOR_LEAD_ID) -> None:
    try:
        from app.api.main import redis_client

        if redis_client is None:
            return

        keys = [
            f"sim_resp:{sessao_id}",
            f"queue:simulador:{sessao_id}",
            f"last_msg:simulador:{sessao_id}",
            f"lock:simulador:{sessao_id}",
            f"followup_ativo:simulador:{sessao_id}",
            f"followup_nivel:simulador:{sessao_id}",
            f"last_msg_time:simulador:{sessao_id}",
        ]

        await redis_client.delete(*keys)

        async for key in redis_client.scan_iter(match=f"processando:simulador:{sessao_id}:*"):
            await redis_client.delete(key)
    except Exception:
        pass


def _lead_id_eh_simulador_ou_invalido(lead_id: str | None) -> bool:
    return str(lead_id or "") == SIMULADOR_LEAD_ID or _parse_uuid_or_none(lead_id) is None


def _normalizar_cor_tag(cor: str | None) -> str:
    cor_limpa = str(cor or "").strip()
    if not cor_limpa:
        return "#2563eb"
    if cor_limpa.startswith("#") and len(cor_limpa) in {4, 7}:
        return cor_limpa
    return "#2563eb"


def _detect_lead_columns(fieldnames: list[str]) -> tuple[str | None, str | None]:
    normalized = {str(name).strip().lower(): name for name in fieldnames if str(name).strip()}

    nome_candidates = ["nome", "nome_contato", "lead", "cliente", "contato"]
    telefone_candidates = ["telefone", "telefone_contato", "whatsapp", "celular", "phone"]

    nome_col = next((normalized[key] for key in nome_candidates if key in normalized), None)
    telefone_col = next((normalized[key] for key in telefone_candidates if key in normalized), None)
    return nome_col, telefone_col


def _load_spreadsheet_rows(file_bytes: bytes, filename: str | None) -> tuple[list[str], list[dict[str, Any]]]:
    nome_arquivo = str(filename or "").lower()

    if nome_arquivo.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Importação de Excel requer a biblioteca openpyxl instalada no servidor.",
            ) from exc

        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return [], []

        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        parsed_rows: list[dict[str, Any]] = []
        for row in rows[1:]:
            parsed_rows.append(
                {
                    headers[index]: ("" if value is None else str(value))
                    for index, value in enumerate(row)
                    if index < len(headers) and headers[index]
                }
            )
        return headers, parsed_rows

    content = file_bytes.decode("utf-8-sig")
    sample = content[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;|\t")
    except Exception:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(content), dialect=dialect)
    return reader.fieldnames or [], list(reader)


async def _garantir_etapa_inicial_crm(db: AsyncSession, empresa_uuid: uuid.UUID) -> uuid.UUID | None:
    result_funil = await db.execute(
        select(CRMFunil)
        .where(CRMFunil.empresa_id == empresa_uuid)
        .options(selectinload(CRMFunil.etapas))
    )
    funil = result_funil.scalars().first()

    if not funil:
        funil = CRMFunil(empresa_id=empresa_uuid, nome="Pipeline Padrão")
        db.add(funil)
        await db.flush()

    etapas = sorted(getattr(funil, "etapas", []) or [], key=lambda etapa: etapa.ordem)
    if etapas:
        return etapas[0].id

    etapa_inicial = CRMEtapa(funil_id=funil.id, nome="Novo Lead", tipo="entrada", ordem=1)
    db.add(etapa_inicial)
    await db.flush()
    return etapa_inicial.id

@router.post("/", response_model=EmpresaResponse, status_code=status.HTTP_201_CREATED)
async def criar_empresa(empresa: EmpresaCreate, db: AsyncSession = Depends(get_db)):
    """
    Cria uma nova Empresa no banco de dados.
    """
    nova_empresa = Empresa(
        nome_empresa=empresa.nome_empresa,
        logo_url=empresa.logo_url,
        credenciais_canais=empresa.credenciais_canais,
        ia_instrucoes_personalizadas=empresa.ia_instrucoes_personalizadas,
        ia_identidade=empresa.ia_identidade,
        ia_regras_negocio=empresa.ia_regras_negocio,
        ia_estrategia_vendas=empresa.ia_estrategia_vendas,
        ia_formatacao_whatsapp=empresa.ia_formatacao_whatsapp,
        ia_tom_voz=empresa.ia_tom_voz,
    )
    db.add(nova_empresa)
    try:
        await db.commit()
        await db.refresh(nova_empresa)
        return nova_empresa
    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Erro ao criar empresa: {str(e)}"
        )


@router.get("/", response_model=List[EmpresaResponse], status_code=status.HTTP_200_OK)
async def listar_empresas(db: AsyncSession = Depends(get_db)):
    """
    Retorna a lista de todas as Empresas cadastradas.
    """
    try:
        result = await db.execute(select(Empresa))
        empresas = result.scalars().all()
        return empresas
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erro ao listar empresas: {str(e)}"
        )


@router.get("/{empresa_id}", response_model=EmpresaResponse, status_code=status.HTTP_200_OK)
async def obter_empresa(empresa_id: str, db: AsyncSession = Depends(get_db)):
    try:
        emp_uuid = uuid.UUID(empresa_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID da empresa inválido")

    result = await db.execute(select(Empresa).where(Empresa.id == emp_uuid))
    empresa = result.scalars().first()
    if not empresa:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Empresa não encontrada")
    # Hardening para bases legadas onde os campos podem estar nulos.
    return {
        "id": empresa.id,
        "nome_empresa": empresa.nome_empresa,
        "area_atuacao": empresa.area_atuacao,
        "logo_url": empresa.logo_url,
        "credenciais_canais": empresa.credenciais_canais or {},
        "ia_instrucoes_personalizadas": empresa.ia_instrucoes_personalizadas,
        "ia_identidade": getattr(empresa, "ia_identidade", None),
        "ia_regras_negocio": getattr(empresa, "ia_regras_negocio", None),
        "ia_estrategia_vendas": getattr(empresa, "ia_estrategia_vendas", None),
        "ia_formatacao_whatsapp": getattr(empresa, "ia_formatacao_whatsapp", None),
        "ia_tom_voz": empresa.ia_tom_voz,
        "conexao_disparo_id": empresa.conexao_disparo_id,
        "disparo_delay_min": empresa.disparo_delay_min if empresa.disparo_delay_min is not None else 3,
        "disparo_delay_max": empresa.disparo_delay_max if empresa.disparo_delay_max is not None else 7,
        "limite_certeza": empresa.limite_certeza if getattr(empresa, "limite_certeza", None) is not None else 0.65,
        "limite_duvida": empresa.limite_duvida if getattr(empresa, "limite_duvida", None) is not None else 0.45,
        "max_agentes_desempate": empresa.max_agentes_desempate if getattr(empresa, "max_agentes_desempate", None) is not None else 3,
    }

# --- ROTAS DA IA CONFIGURATOR ---
from fastapi import Header

async def require_ia_config_access(
    empresa_id: str,
    db: AsyncSession = Depends(get_db),
    x_user_id: Optional[str] = Header(None),
):
    if not x_user_id:
        raise HTTPException(status_code=401, detail="Usuário não autenticado.")

    try:
        user_uuid = uuid.UUID(x_user_id)
    except ValueError:
        raise HTTPException(status_code=401, detail="Identificador de usuário inválido.")

    result = await db.execute(select(Usuario).where(Usuario.id == user_uuid))
    usuario_bd = result.scalars().first()
    if not usuario_bd:
        raise HTTPException(status_code=401, detail="Usuário não encontrado.")

    if is_root_admin_email(usuario_bd.email) or is_super_admin_role(usuario_bd.role):
        return usuario_bd

    print(f"Role no Banco: {usuario_bd.role}")
    if is_admin_empresa_role(usuario_bd.role):
        try:
            empresa_uuid = uuid.UUID(empresa_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="ID da empresa inválido.")

        if usuario_bd.empresa_id != empresa_uuid:
            raise HTTPException(status_code=403, detail="Acesso negado para esta empresa.")
        return usuario_bd

    raise HTTPException(
        status_code=403,
        detail="Acesso negado. Apenas Super Admin ou Admin da própria empresa."
    )

@router.get("/{empresa_id}/ia-config", response_model=IAConfigResponse, status_code=status.HTTP_200_OK)
async def get_ia_config(
    empresa_id: str,
    db: AsyncSession = Depends(get_db),
    _: Usuario = Depends(require_ia_config_access),
):
    result = await db.execute(select(Empresa).where(Empresa.id == empresa_id))
    empresa = result.scalars().first()
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    return {
        "ia_instrucoes_personalizadas": empresa.ia_instrucoes_personalizadas, 
        "ia_identidade": getattr(empresa, "ia_identidade", None),
        "ia_regras_negocio": getattr(empresa, "ia_regras_negocio", None),
        "ia_estrategia_vendas": getattr(empresa, "ia_estrategia_vendas", None),
        "ia_formatacao_whatsapp": getattr(empresa, "ia_formatacao_whatsapp", None),
        "ia_tom_voz": empresa.ia_tom_voz,
        "nome_agente": empresa.nome_agente,
        "mensagem_saudacao": empresa.mensagem_saudacao,
        "modelo_ia": empresa.modelo_ia,
        "modelo_roteador": getattr(empresa, 'modelo_roteador', 'gpt-4o-mini'),
        "followup_ativo": getattr(empresa, 'followup_ativo', False) or False,
        "followup_espera_nivel_1_minutos": getattr(empresa, 'followup_espera_nivel_1_minutos', 20) or 20,
        "followup_espera_nivel_2_minutos": getattr(empresa, 'followup_espera_nivel_2_minutos', 10) or 10,
        "limite_certeza": getattr(empresa, "limite_certeza", 0.65) if getattr(empresa, "limite_certeza", None) is not None else 0.65,
        "limite_duvida": getattr(empresa, "limite_duvida", 0.45) if getattr(empresa, "limite_duvida", None) is not None else 0.45,
        "max_agentes_desempate": getattr(empresa, "max_agentes_desempate", 3) if getattr(empresa, "max_agentes_desempate", None) is not None else 3,
        "informacoes_adicionais": getattr(empresa, 'informacoes_adicionais', None),
        "coletar_nome": getattr(empresa, 'coletar_nome', True) if getattr(empresa, 'coletar_nome', True) is not None else True,
    }

@router.put("/{empresa_id}/ia-config", response_model=IAConfigResponse, status_code=status.HTTP_200_OK)
@router.post("/{empresa_id}/ia-config", response_model=IAConfigResponse, status_code=status.HTTP_200_OK)
async def put_ia_config(
    empresa_id: str,
    data: IAConfigUpdateRequest,
    db: AsyncSession = Depends(get_db),
    _: Usuario = Depends(require_ia_config_access),
):
    # Deprecated para agenda: horários de funcionamento devem ser atualizados em PUT /empresas/{empresa_id}/agenda.
    result = await db.execute(select(Empresa).where(Empresa.id == empresa_id))
    empresa = result.scalars().first()
    if not empresa:
        raise HTTPException(status_code=404, detail="Empresa não encontrada")
    if data.ia_instrucoes_personalizadas is not None:
        empresa.ia_instrucoes_personalizadas = data.ia_instrucoes_personalizadas
    if data.ia_identidade is not None:
        empresa.ia_identidade = data.ia_identidade
    if data.ia_regras_negocio is not None:
        empresa.ia_regras_negocio = data.ia_regras_negocio
    if data.ia_estrategia_vendas is not None:
        empresa.ia_estrategia_vendas = data.ia_estrategia_vendas
    if data.ia_formatacao_whatsapp is not None:
        empresa.ia_formatacao_whatsapp = data.ia_formatacao_whatsapp
    if data.ia_tom_voz is not None:
        empresa.ia_tom_voz = data.ia_tom_voz
    if data.nome_agente is not None:
        empresa.nome_agente = data.nome_agente
    if data.mensagem_saudacao is not None:
        empresa.mensagem_saudacao = data.mensagem_saudacao
    if data.modelo_ia is not None:
        empresa.modelo_ia = data.modelo_ia
    if data.modelo_roteador is not None:
        empresa.modelo_roteador = data.modelo_roteador
    if data.followup_ativo is not None:
        empresa.followup_ativo = data.followup_ativo
    if data.followup_espera_nivel_1_minutos is not None:
        empresa.followup_espera_nivel_1_minutos = data.followup_espera_nivel_1_minutos
    if data.followup_espera_nivel_2_minutos is not None:
        empresa.followup_espera_nivel_2_minutos = data.followup_espera_nivel_2_minutos
    if data.limite_certeza is not None:
        empresa.limite_certeza = data.limite_certeza
    if data.limite_duvida is not None:
        empresa.limite_duvida = data.limite_duvida
    if data.max_agentes_desempate is not None:
        empresa.max_agentes_desempate = data.max_agentes_desempate
    if data.informacoes_adicionais is not None:
        empresa.informacoes_adicionais = data.informacoes_adicionais
    if data.coletar_nome is not None:
        empresa.coletar_nome = data.coletar_nome
    limite_duvida = empresa.limite_duvida if getattr(empresa, "limite_duvida", None) is not None else 0.45
    limite_certeza = empresa.limite_certeza if getattr(empresa, "limite_certeza", None) is not None else 0.65
    max_desempate = empresa.max_agentes_desempate if getattr(empresa, "max_agentes_desempate", None) is not None else 3
    if limite_duvida < 0 or limite_duvida > 1:
        raise HTTPException(status_code=400, detail="limite_duvida deve estar entre 0 e 1")
    if limite_certeza < 0 or limite_certeza > 1:
        raise HTTPException(status_code=400, detail="limite_certeza deve estar entre 0 e 1")
    if limite_duvida > limite_certeza:
        raise HTTPException(status_code=400, detail="limite_duvida não pode ser maior que limite_certeza")
    if max_desempate < 1:
        raise HTTPException(status_code=400, detail="max_agentes_desempate deve ser >= 1")
    try:
        await db.commit()
        await db.refresh(empresa)
        return {
            "ia_instrucoes_personalizadas": empresa.ia_instrucoes_personalizadas, 
            "ia_identidade": getattr(empresa, "ia_identidade", None),
            "ia_regras_negocio": getattr(empresa, "ia_regras_negocio", None),
            "ia_estrategia_vendas": getattr(empresa, "ia_estrategia_vendas", None),
            "ia_formatacao_whatsapp": getattr(empresa, "ia_formatacao_whatsapp", None),
            "ia_tom_voz": empresa.ia_tom_voz,
            "nome_agente": empresa.nome_agente,
            "mensagem_saudacao": empresa.mensagem_saudacao,
            "modelo_ia": empresa.modelo_ia,
            "modelo_roteador": getattr(empresa, 'modelo_roteador', 'gpt-4o-mini'),
            "followup_ativo": getattr(empresa, 'followup_ativo', False) or False,
            "followup_espera_nivel_1_minutos": getattr(empresa, 'followup_espera_nivel_1_minutos', 20) or 20,
            "followup_espera_nivel_2_minutos": getattr(empresa, 'followup_espera_nivel_2_minutos', 10) or 10,
            "limite_certeza": getattr(empresa, "limite_certeza", 0.65) if getattr(empresa, "limite_certeza", None) is not None else 0.65,
            "limite_duvida": getattr(empresa, "limite_duvida", 0.45) if getattr(empresa, "limite_duvida", None) is not None else 0.45,
            "max_agentes_desempate": getattr(empresa, "max_agentes_desempate", 3) if getattr(empresa, "max_agentes_desempate", None) is not None else 3,
            "informacoes_adicionais": getattr(empresa, 'informacoes_adicionais', None),
            "coletar_nome": getattr(empresa, 'coletar_nome', True) if getattr(empresa, 'coletar_nome', True) is not None else True,
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao atualizar: {str(e)}")

@router.put("/{empresa_id}", response_model=EmpresaResponse, status_code=status.HTTP_200_OK)
async def atualizar_empresa(empresa_id: str, data: EmpresaUpdate, db: AsyncSession = Depends(get_db)):
    """
    Atualiza dados básicos da Empresa (nome e área de atuação).
    """
    result = await db.execute(select(Empresa).where(Empresa.id == empresa_id))
    empresa = result.scalars().first()
    if not empresa:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Empresa não encontrada")
    
    if data.nome_empresa is not None:
        empresa.nome_empresa = data.nome_empresa
    if data.area_atuacao is not None:
        empresa.area_atuacao = data.area_atuacao
    if data.logo_url is not None:
        empresa.logo_url = data.logo_url
    if data.ia_instrucoes_personalizadas is not None:
        empresa.ia_instrucoes_personalizadas = data.ia_instrucoes_personalizadas
    if data.ia_identidade is not None:
        empresa.ia_identidade = data.ia_identidade
    if data.ia_regras_negocio is not None:
        empresa.ia_regras_negocio = data.ia_regras_negocio
    if data.ia_estrategia_vendas is not None:
        empresa.ia_estrategia_vendas = data.ia_estrategia_vendas
    if data.ia_formatacao_whatsapp is not None:
        empresa.ia_formatacao_whatsapp = data.ia_formatacao_whatsapp
    if data.ia_tom_voz is not None:
        empresa.ia_tom_voz = data.ia_tom_voz
    if data.disparo_delay_min is not None:
        if data.disparo_delay_min < 0:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="disparo_delay_min não pode ser negativo")
        empresa.disparo_delay_min = data.disparo_delay_min
    if data.disparo_delay_max is not None:
        if data.disparo_delay_max < 0:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="disparo_delay_max não pode ser negativo")
        empresa.disparo_delay_max = data.disparo_delay_max
    if data.limite_certeza is not None:
        if data.limite_certeza < 0 or data.limite_certeza > 1:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="limite_certeza deve estar entre 0 e 1")
        empresa.limite_certeza = data.limite_certeza
    if data.limite_duvida is not None:
        if data.limite_duvida < 0 or data.limite_duvida > 1:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="limite_duvida deve estar entre 0 e 1")
        empresa.limite_duvida = data.limite_duvida
    if data.max_agentes_desempate is not None:
        if data.max_agentes_desempate < 1:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="max_agentes_desempate deve ser >= 1")
        empresa.max_agentes_desempate = data.max_agentes_desempate
    if data.conexao_disparo_id is not None:
        if data.conexao_disparo_id == "":
            empresa.conexao_disparo_id = None
        else:
            try:
                conexao_disparo_uuid = uuid.UUID(data.conexao_disparo_id)
            except ValueError:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="conexao_disparo_id inválido")

            result_conexao = await db.execute(
                select(Conexao).where(
                    Conexao.id == conexao_disparo_uuid,
                    Conexao.empresa_id == empresa.id,
                )
            )
            conexao = result_conexao.scalars().first()
            if not conexao:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Conexão de disparo não encontrada para esta empresa",
                )
            empresa.conexao_disparo_id = conexao.id

    delay_min = empresa.disparo_delay_min if empresa.disparo_delay_min is not None else 3
    delay_max = empresa.disparo_delay_max if empresa.disparo_delay_max is not None else 7
    if delay_min > delay_max:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="O intervalo mínimo de disparo não pode ser maior que o máximo",
        )
    limite_duvida = empresa.limite_duvida if getattr(empresa, "limite_duvida", None) is not None else 0.45
    limite_certeza = empresa.limite_certeza if getattr(empresa, "limite_certeza", None) is not None else 0.65
    if limite_duvida > limite_certeza:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="limite_duvida não pode ser maior que limite_certeza",
        )
        
    try:
        await db.commit()
        await db.refresh(empresa)
        return empresa
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Erro ao atualizar empresa: {str(e)}")

@router.delete("/{empresa_id}", status_code=status.HTTP_204_NO_CONTENT)
async def deletar_empresa(empresa_id: str, db: AsyncSession = Depends(get_db)):
    """
    Remove uma Empresa e todos os dados vinculados (via cascade configurado no banco).
    """
    import uuid
    try:
        emp_uuid = uuid.UUID(empresa_id)
    except ValueError:
         raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID da empresa inválido")
         
    result = await db.execute(select(Empresa).where(Empresa.id == emp_uuid))
    empresa = result.scalars().first()
    if not empresa:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Empresa não encontrada")
        
    try:
        await db.delete(empresa)
        await db.commit()
        return None
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Erro ao deletar empresa: {str(e)}")

@router.post("/setup", status_code=status.HTTP_201_CREATED)
async def setup_empresa(data: EmpresaSetupRequest, db: AsyncSession = Depends(get_db)):
    """
    Cria uma nova Empresa e seu respectivo Usuário admin (admin_empresa).
    """
    print(f"Iniciando setup_empresa para: {data.nome_empresa}")
    
    # Verifica se o email já existe
    admin_email_normalizado = normalize_user_email(data.admin_email)
    result = await db.execute(select(Usuario).where(Usuario.email == admin_email_normalizado))
    if result.scalars().first():
        print(f"Setup falhou: E-mail {data.admin_email} já em uso.")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="E-mail já está em uso por outro usuário."
        )

    try:
        print("Criando empresa...")
        nova_empresa = Empresa(
            nome_empresa=data.nome_empresa,
            area_atuacao=data.area_atuacao,
            logo_url=data.logo_url,
            ia_identidade=data.ia_identidade,
            ia_regras_negocio=data.ia_regras_negocio,
            ia_estrategia_vendas=data.ia_estrategia_vendas,
            ia_formatacao_whatsapp=data.ia_formatacao_whatsapp,
        )
        db.add(nova_empresa)
        await db.flush() # Para obter o ID da empresa gerado
        
        print("Criando usuário admin...")
        is_root_admin = is_root_admin_email(data.admin_email)
        novo_usuario = Usuario(
            empresa_id=None if is_root_admin else nova_empresa.id,
            nome=data.admin_nome,
            email=admin_email_normalizado,
            senha_hash=get_password_hash(data.admin_senha),
            role=ROOT_ADMIN_ROLE if is_root_admin else ADMIN_EMPRESA_ROLE,
            ativo=True
        )
        db.add(novo_usuario)
        
        print("Commitando transação...")
        await db.commit()
        await db.refresh(nova_empresa)
        await db.refresh(novo_usuario)
        
        print(f"Setup concluído com sucesso. ID da empresa: {nova_empresa.id}")
        return {
            "mensagem": "Empresa e usuário criados com sucesso!",
            "empresa": {
                "id": str(nova_empresa.id),
                "nome": nova_empresa.nome_empresa
            },
            "usuario": {
                "id": str(novo_usuario.id),
                "email": novo_usuario.email
            }
        }
    except ValueError as e:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        print(f"Erro durante o setup da empresa: {str(e)}")
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erro banco de dados: {str(e)}"
        )


@router.get("/{empresa_id}/dashboard")
async def get_dashboard_data(empresa_id: str):
    """
    Retorna dados mockados para o Dashboard do Cliente (Tenant).
    """
    return {
        "total_leads": 12,
        "agendamentos_hoje": 3,
        "status_ia": "Ativa"
    }


class RAGCreateRequest(BaseModel):
    tipo: str
    conteudo: str


class RAGDeleteRequest(BaseModel):
    source_name: str


class CRMLeadResponse(BaseModel):
    id: str
    nome_contato: str
    telefone: str | None = None
    historico_resumo: str | None = None
    tags: List[str] = Field(default_factory=list)
    dados_adicionais: Dict[str, Any] = Field(default_factory=dict)
    valor_conversao: float | None = None
    criado_em: str | None = None


class CRMEtapaResponse(BaseModel):
    id: str
    nome: str
    tipo: str | None = None
    ordem: int
    leads: List[CRMLeadResponse]


class CRMFunilResponse(BaseModel):
    funil_id: str
    funil_nome: str
    etapas: List[CRMEtapaResponse]


class CRMEtapaUpdateRequest(BaseModel):
    nome: str | None = None
    tipo: str | None = None
    ordem: int | None = None


class CRMLeadUpdateRequest(BaseModel):
    nome_contato: str | None = None
    telefone_contato: str | None = None
    historico_resumo: str | None = None
    etapa_id: str | None = None
    status_atendimento: str | None = None
    tags: List[str] | None = None
    dados_adicionais: Dict[str, Any] | None = None
    valor_conversao: float | None = None


class DestinoTransferenciaBase(BaseModel):
    nome_destino: str
    contatos_destino: List[str] = Field(default_factory=list)
    instrucoes_ativacao: str | None = None


class DestinoTransferenciaCreate(DestinoTransferenciaBase):
    pass


class DestinoTransferenciaUpdate(BaseModel):
    nome_destino: str | None = None
    contatos_destino: List[str] | None = None
    instrucoes_ativacao: str | None = None


class DestinoTransferenciaResponse(DestinoTransferenciaBase):
    id: str
    criado_em: str | None = None


class HistoricoTransferenciaResponse(BaseModel):
    id: str
    criado_em: str | None = None
    lead_id: str
    lead_nome: str | None = None
    destino_id: str | None = None
    destino_nome: str | None = None
    motivo_ia: str | None = None
    resumo_enviado: str | None = None


class TransferenciaManualRequest(BaseModel):
    destino_id: str


class TagGroupBase(BaseModel):
    nome: str
    cor: str | None = None
    ordem: int = 0


class TagGroupCreate(TagGroupBase):
    pass


class TagGroupUpdate(BaseModel):
    nome: str | None = None
    cor: str | None = None
    ordem: int | None = None


class TagGroupResponse(TagGroupBase):
    id: str


class TagCRMBase(BaseModel):
    nome: str
    cor: str = "#2563eb"
    instrucao_ia: str | None = None
    grupo_id: str | None = None
    disparar_conversao_ads: bool = False


class TagCRMCreate(TagCRMBase):
    pass


class TagCRMUpdate(BaseModel):
    nome: str | None = None
    cor: str | None = None
    instrucao_ia: str | None = None
    grupo_id: str | None = None
    disparar_conversao_ads: bool | None = None


class TagCRMResponse(TagCRMBase):
    id: str
    criado_em: str | None = None


class CampanhaListaResponse(BaseModel):
    tag_id: str
    tag: str
    cor: str
    total_leads: int


class CampanhaListaLeadResponse(BaseModel):
    id: str
    nome_contato: str
    telefone_contato: str | None = None
    status: str | None = None
    tags: List[str] = Field(default_factory=list)
    historico_resumo: str | None = None
    dados_adicionais: Dict[str, Any] = Field(default_factory=dict)


class TemplateMensagemBase(BaseModel):
    nome: str
    texto_template: str


class TemplateMensagemCreate(TemplateMensagemBase):
    variaveis_esperadas: List[str] | None = None


class TemplateMensagemUpdate(BaseModel):
    nome: str | None = None
    texto_template: str | None = None
    variaveis_esperadas: List[str] | None = None


class TemplateMensagemResponse(BaseModel):
    id: str
    nome: str
    texto_template: str
    variaveis_esperadas: List[str] = Field(default_factory=list)
    criado_em: str | None = None


class CampanhaDisparoCreate(BaseModel):
    nome: str
    template_id: str
    tags_alvo: List[str] = Field(default_factory=list)
    data_agendamento: str | None = None


class CampanhaDisparoResponse(BaseModel):
    id: str
    nome: str
    template_id: str | None = None
    template_nome: str | None = None
    tags_alvo: List[str] = Field(default_factory=list)
    data_agendamento: str | None = None
    status: str
    criado_em: str | None = None


class CampanhaPreviewResponse(BaseModel):
    preview_texto: str
    total_leads: int
    usou_mock: bool

@router.get("/{empresa_id}/rag")
async def listar_conhecimento_rag(
    empresa_id: str,
    db: AsyncSession = Depends(get_db),
    _: Usuario = Depends(require_ia_config_access),
):
    """
    Retorna a lista de documentos lógicos da base vetorial da empresa.
    """
    try:
        result = await db.execute(
            select(
                Conhecimento.source_name.label("source_name"),
                Conhecimento.source_type.label("source_type"),
                func.count(Conhecimento.id).label("total_chunks"),
                func.min(Conhecimento.criado_em).label("criado_em"),
            )
            .where(
                Conhecimento.empresa_id == empresa_id,
                Conhecimento.source_name.isnot(None),
                func.length(func.trim(Conhecimento.source_name)) > 0,
            )
            .group_by(Conhecimento.source_name, Conhecimento.source_type)
            .order_by(func.min(Conhecimento.criado_em).desc())
        )
        documentos = result.all()
        return [
            {
                "source_name": row.source_name,
                "source_type": row.source_type,
                "total_chunks": int(row.total_chunks or 0),
                "criado_em": row.criado_em.isoformat() if row.criado_em else None,
            }
            for row in documentos
        ]
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erro ao buscar conhecimento RAG: {str(e)}"
        )

@router.post("/{empresa_id}/rag", status_code=status.HTTP_201_CREATED)
async def adicionar_conhecimento_rag(
    empresa_id: str,
    data: RAGCreateRequest,
    db: AsyncSession = Depends(get_db),
    _: Usuario = Depends(require_ia_config_access),
):
    """
    Adiciona um novo conhecimento RAG (texto ou url) processando com Langchain e embeddings, igual ao PDF.
    """
    if data.tipo not in ["texto", "url", "pdf"]:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Tipo de conhecimento inválido.")

    texto_para_processar = ""
    label_fonte = ""
    source_name = ""

    if data.tipo == "texto":
        texto_para_processar = data.conteudo
        label_fonte = "[Texto Manual]"
        source_name = "Texto Manual"
    elif data.tipo == "url":
        try:
            import requests
            from bs4 import BeautifulSoup
            
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
            response = requests.get(data.conteudo, headers=headers, timeout=10)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.text, 'html.parser')
            # Remove scripts, styles, etc
            for script in soup(["script", "style", "header", "footer", "nav"]):
                script.extract()
                
            texto_para_processar = soup.get_text(separator=' ', strip=True)
            label_fonte = f"[{data.conteudo}]"
            source_name = data.conteudo
            
            if not texto_para_processar:
                 raise ValueError("Nenhum texto principal encontrado na URL.")
                 
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Erro ao extrair conteúdo da URL: {str(e)}")

    if not texto_para_processar.strip():
        raise HTTPException(status_code=400, detail="O conteúdo fornecido está vazio.")

    try:
        from langchain_text_splitters import RecursiveCharacterTextSplitter
        text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=1000,
            chunk_overlap=200,
            length_function=len,
        )
        chunks = text_splitter.split_text(texto_para_processar)
        
        from langchain_openai import OpenAIEmbeddings
        embeddings_model = OpenAIEmbeddings(model="text-embedding-3-small")
        chunks_embeddings = await embeddings_model.aembed_documents(chunks)
        
        primeiro_id = None
        for i, (chunk, emb) in enumerate(zip(chunks, chunks_embeddings)):
            novo_chunk_rag = ConhecimentoRAG(
                empresa_id=empresa_id,
                tipo=data.tipo,
                conteudo=f"{label_fonte} " + chunk,
                source_name=source_name,
                source_type=data.tipo,
                status="Ativo"
            )
            db.add(novo_chunk_rag)
            await db.flush() # obtem o id
            
            if i == 0:
                primeiro_id = novo_chunk_rag.id
            
            novo_vetor = Conhecimento(
                empresa_id=empresa_id,
                conteudo=chunk,
                source_name=source_name,
                source_type=data.tipo,
                embedding=emb
            )
            db.add(novo_vetor)
            
        await db.commit()
        return {"mensagem": "Conhecimento adicionado com sucesso!", "id": str(primeiro_id)}
        
    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erro no processamento de RAG: {str(e)}"
        )


async def processar_rag_pdf_background(empresa_id: str, content: bytes, filename: str):
    try:
        async with AsyncSessionLocal() as db:
            source_name = filename or "arquivo.pdf"
            pdf_stream = io.BytesIO(content)
            text_content = ""
            try:
                with pdfplumber.open(pdf_stream) as pdf:
                    num_pages = len(pdf.pages)
                    print(f"Lendo PDF com pdfplumber: {source_name} ({num_pages} páginas)")
                    for page in pdf.pages:
                        page_text = page.extract_text()
                        if page_text:
                            text_content += page_text + "\n"
            except Exception as e:
                print(f"Erro ao abrir PDF com pdfplumber: {str(e)}")
                return

            if len(text_content.strip()) < 10:
                print(f"Aviso: PDF {source_name} extraiu pouco ou nenhum texto. Verifique se é uma imagem.")
                # Se for imagem, o ideal futuramente seria OCR, por hora apenas logamos.
                return

            from langchain_text_splitters import RecursiveCharacterTextSplitter
            text_splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=200, length_function=len)
            chunks = text_splitter.split_text(text_content)

            from langchain_openai import OpenAIEmbeddings
            embeddings_model = OpenAIEmbeddings(model="text-embedding-3-small")
            chunks_embeddings = await embeddings_model.aembed_documents(chunks)

            for chunk, emb in zip(chunks, chunks_embeddings):
                novo_chunk_rag = ConhecimentoRAG(
                    empresa_id=empresa_id,
                    tipo="pdf",
                    conteudo=f"[{filename}] " + chunk,
                    source_name=filename or "arquivo.pdf",
                    source_type="pdf",
                    status="Ativo"
                )
                db.add(novo_chunk_rag)

                novo_vetor = Conhecimento(
                    empresa_id=empresa_id,
                    conteudo=chunk,
                    source_name=filename or "arquivo.pdf",
                    source_type="pdf",
                    embedding=emb
                )
                db.add(novo_vetor)

            await db.commit()
            print(f"Sucesso: PDF {filename} processado!")
    except Exception as e:
        traceback.print_exc()
        print(f"Erro no processamento background do PDF {filename}: {str(e)}")


@router.post("/{empresa_id}/rag/pdf", status_code=status.HTTP_202_ACCEPTED)
async def adicionar_conhecimento_rag_pdf(
    empresa_id: str,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    _: Usuario = Depends(require_ia_config_access),
):
    """
    Cadastra e vetoriza um arquivo PDF na base de conhecimento.
    """
    if file.content_type != "application/pdf":
        raise HTTPException(status_code=400, detail="Arquivo deve ser um PDF.")
        
    content = await file.read()
    background_tasks.add_task(processar_rag_pdf_background, empresa_id, content, file.filename)
    return {"mensagem": "Upload concluído. O arquivo está sendo processado em segundo plano e aparecerá na lista em breve."}


@router.delete("/{empresa_id}/rag")
async def deletar_conhecimento_rag_por_source_name(
    empresa_id: str,
    payload: RAGDeleteRequest,
    db: AsyncSession = Depends(get_db),
    _: Usuario = Depends(require_ia_config_access),
):
    """
    Exclui um documento lógico da base RAG pelo source_name da empresa.
    Remove chunks vetoriais (conhecimento) e chunks de visualização (conhecimento_rag).
    """
    source_name = str(payload.source_name or "").strip()
    if not source_name:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="source_name é obrigatório.",
        )

    try:
        resultado_conhecimento = await db.execute(
            delete(Conhecimento).where(
                Conhecimento.empresa_id == empresa_id,
                Conhecimento.source_name == source_name,
            )
        )
        resultado_rag = await db.execute(
            delete(ConhecimentoRAG).where(
                ConhecimentoRAG.empresa_id == empresa_id,
                or_(
                    ConhecimentoRAG.source_name == source_name,
                    ConhecimentoRAG.conteudo.like(f"[{source_name}] %"),
                ),
            )
        )
        await db.commit()

        return {
            "mensagem": "Documento RAG removido com sucesso.",
            "source_name": source_name,
            "registros_conhecimento_removidos": int(resultado_conhecimento.rowcount or 0),
            "registros_conhecimento_rag_removidos": int(resultado_rag.rowcount or 0),
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erro ao excluir conhecimento RAG: {str(e)}",
        )


@router.get("/{empresa_id}/crm", response_model=CRMFunilResponse)
async def obter_crm(empresa_id: str, db: AsyncSession = Depends(get_db)):
    """
    Retorna a estrutura do Funil da empresa com as Etapas e os Leads aninhados.
    Se não existir um funil, cria um padrão ("Pipeline Padrão" com colunas: "Novo Lead", "Em Atendimento", "Fechado").
    """
    try:
        # Busca o primeiro funil da empresa, carregando as etapas e os leads das etapas
        result = await db.execute(
            select(CRMFunil)
            .where(CRMFunil.empresa_id == empresa_id)
            .options(
                selectinload(CRMFunil.etapas).selectinload(CRMEtapa.leads)
            )
        )
        funil = result.scalars().first()

        # Se não existir, cria o funil padrão
        if not funil:
            novo_funil = CRMFunil(empresa_id=empresa_id, nome="Pipeline Padrão")
            db.add(novo_funil)
            await db.flush() # Para pegar o ID do funil

            etapas_padrao = [
                CRMEtapa(funil_id=novo_funil.id, nome="Novo Lead", tipo="entrada", ordem=1),
                CRMEtapa(funil_id=novo_funil.id, nome="Em Atendimento", tipo="atendimento", ordem=2),
                CRMEtapa(funil_id=novo_funil.id, nome="Fechado", tipo="fechamento", ordem=3)
            ]
            db.add_all(etapas_padrao)
            await db.commit()
            
            # Recarrega o funil agora com as etapas
            result = await db.execute(
                select(CRMFunil)
                .where(CRMFunil.id == novo_funil.id)
                .options(
                    selectinload(CRMFunil.etapas).selectinload(CRMEtapa.leads)
                )
            )
            funil = result.scalars().first()

        # Ordena as etapas pela ordem
        etapas_ordenadas = sorted(funil.etapas, key=lambda x: x.ordem)

        resposta: dict = {
            "funil_id": str(funil.id),
            "funil_nome": funil.nome,
            "etapas": []
        }

        for etapa in etapas_ordenadas:
            etapa_dict = {
                "id": str(etapa.id),
                "nome": etapa.nome,
                "tipo": etapa.tipo,
                "ordem": etapa.ordem,
                "leads": [
                    {
                        "id": str(lead.id),
                        "nome_contato": lead.nome_contato,
                        "telefone": lead.telefone_contato,
                        "historico_resumo": lead.historico_resumo,
                        "tags": lead.tags or [],
                        "dados_adicionais": lead.dados_adicionais or {},
                        "valor_conversao": lead.valor_conversao,
                        "criado_em": lead.criado_em.isoformat() if lead.criado_em else None
                    } for lead in etapa.leads
                ]
            }
            resposta["etapas"].append(etapa_dict)

        return resposta

    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erro ao carregar CRM: {str(e)}"
        )


@router.get("/{empresa_id}/transferencias/destinos", response_model=List[DestinoTransferenciaResponse])
async def listar_destinos_transferencia(empresa_id: str, db: AsyncSession = Depends(get_db)):
    try:
        emp_uuid = uuid.UUID(empresa_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID de empresa inválido")

    result = await db.execute(
        select(DestinosTransferencia)
        .where(DestinosTransferencia.empresa_id == emp_uuid)
        .order_by(DestinosTransferencia.criado_em.desc())
    )
    destinos = result.scalars().all()
    return [
        {
            "id": str(destino.id),
            "nome_destino": destino.nome_destino,
            "contatos_destino": destino.contatos_destino or [],
            "instrucoes_ativacao": destino.instrucoes_ativacao,
            "criado_em": destino.criado_em.isoformat() if destino.criado_em else None,
        }
        for destino in destinos
    ]


@router.post("/{empresa_id}/transferencias/destinos", response_model=DestinoTransferenciaResponse, status_code=status.HTTP_201_CREATED)
async def criar_destino_transferencia(
    empresa_id: str,
    data: DestinoTransferenciaCreate,
    db: AsyncSession = Depends(get_db),
):
    try:
        emp_uuid = uuid.UUID(empresa_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID de empresa inválido")

    contatos_normalizados = [str(contato).strip() for contato in data.contatos_destino if str(contato).strip()]
    destino = DestinosTransferencia(
        empresa_id=emp_uuid,
        nome_destino=data.nome_destino.strip(),
        contatos_destino=contatos_normalizados,
        instrucoes_ativacao=(data.instrucoes_ativacao or "").strip() or None,
    )
    db.add(destino)

    try:
        await db.commit()
        await db.refresh(destino)
        return {
            "id": str(destino.id),
            "nome_destino": destino.nome_destino,
            "contatos_destino": destino.contatos_destino or [],
            "instrucoes_ativacao": destino.instrucoes_ativacao,
            "criado_em": destino.criado_em.isoformat() if destino.criado_em else None,
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Erro ao criar destino: {str(e)}")


@router.put("/{empresa_id}/transferencias/destinos/{destino_id}", response_model=DestinoTransferenciaResponse)
async def atualizar_destino_transferencia(
    empresa_id: str,
    destino_id: str,
    data: DestinoTransferenciaUpdate,
    db: AsyncSession = Depends(get_db),
):
    try:
        emp_uuid = uuid.UUID(empresa_id)
        destino_uuid = uuid.UUID(destino_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID inválido")

    result = await db.execute(
        select(DestinosTransferencia).where(
            DestinosTransferencia.id == destino_uuid,
            DestinosTransferencia.empresa_id == emp_uuid,
        )
    )
    destino = result.scalars().first()
    if not destino:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Destino não encontrado")

    if data.nome_destino is not None:
        destino.nome_destino = data.nome_destino.strip()
    if data.contatos_destino is not None:
        destino.contatos_destino = [str(contato).strip() for contato in data.contatos_destino if str(contato).strip()]
    if data.instrucoes_ativacao is not None:
        destino.instrucoes_ativacao = data.instrucoes_ativacao.strip() or None

    try:
        await db.commit()
        await db.refresh(destino)
        return {
            "id": str(destino.id),
            "nome_destino": destino.nome_destino,
            "contatos_destino": destino.contatos_destino or [],
            "instrucoes_ativacao": destino.instrucoes_ativacao,
            "criado_em": destino.criado_em.isoformat() if destino.criado_em else None,
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Erro ao atualizar destino: {str(e)}")


@router.delete("/{empresa_id}/transferencias/destinos/{destino_id}", status_code=status.HTTP_204_NO_CONTENT)
async def excluir_destino_transferencia(
    empresa_id: str,
    destino_id: str,
    db: AsyncSession = Depends(get_db),
):
    try:
        emp_uuid = uuid.UUID(empresa_id)
        destino_uuid = uuid.UUID(destino_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID inválido")

    result = await db.execute(
        select(DestinosTransferencia).where(
            DestinosTransferencia.id == destino_uuid,
            DestinosTransferencia.empresa_id == emp_uuid,
        )
    )
    destino = result.scalars().first()
    if not destino:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Destino não encontrado")

    try:
        await db.delete(destino)
        await db.commit()
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Erro ao excluir destino: {str(e)}")


@router.post("/{empresa_id}/transferencias/destinos/{destino_id}/testar", status_code=status.HTTP_200_OK)
async def testar_destino_transferencia_endpoint(
    empresa_id: str,
    destino_id: str,
):
    resultado = await testar_destino_transferencia(
        empresa_id=empresa_id,
        destino_id=destino_id,
    )

    if not resultado.get("success"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=resultado.get("detail") or "Falha ao testar destino de transferência.",
        )

    return resultado


@router.get("/{empresa_id}/transferencias/historico", response_model=List[HistoricoTransferenciaResponse])
async def listar_historico_transferencia(empresa_id: str, db: AsyncSession = Depends(get_db)):
    try:
        emp_uuid = uuid.UUID(empresa_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID de empresa inválido")

    result = await db.execute(
        select(HistoricoTransferencia)
        .where(HistoricoTransferencia.empresa_id == emp_uuid)
        .options(
            selectinload(HistoricoTransferencia.lead),
            selectinload(HistoricoTransferencia.destino),
        )
        .order_by(HistoricoTransferencia.criado_em.desc())
    )
    historicos = result.scalars().all()
    return [
        {
            "id": str(item.id),
            "criado_em": item.criado_em.isoformat() if item.criado_em else None,
            "lead_id": str(item.lead_id),
            "lead_nome": item.lead.nome_contato if item.lead else None,
            "destino_id": str(item.destino_id) if item.destino_id else None,
            "destino_nome": item.destino.nome_destino if item.destino else None,
            "motivo_ia": item.motivo_ia,
            "resumo_enviado": item.resumo_enviado,
        }
        for item in historicos
    ]


@router.get("/{empresa_id}/crm/tags/grupos", response_model=List[TagGroupResponse])
async def listar_tags_grupos(empresa_id: str, db: AsyncSession = Depends(get_db)):
    try:
        emp_uuid = uuid.UUID(empresa_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID de empresa inválido")

    result = await db.execute(
        select(TagGroup)
        .where(TagGroup.empresa_id == emp_uuid)
        .order_by(TagGroup.ordem.asc(), TagGroup.nome.asc())
    )
    grupos = result.scalars().all()
    return [
        {
            "id": str(grupo.id),
            "nome": grupo.nome,
            "cor": grupo.cor,
            "ordem": grupo.ordem,
        }
        for grupo in grupos
    ]


@router.post("/{empresa_id}/crm/tags/grupos", response_model=TagGroupResponse, status_code=status.HTTP_201_CREATED)
async def criar_tag_grupo(empresa_id: str, data: TagGroupCreate, db: AsyncSession = Depends(get_db)):
    try:
        emp_uuid = uuid.UUID(empresa_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID de empresa inválido")

    nome_limpo = str(data.nome or "").strip()
    if not nome_limpo:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Nome do grupo é obrigatório")

    result = await db.execute(
        select(TagGroup).where(
            TagGroup.empresa_id == emp_uuid,
            TagGroup.nome.ilike(nome_limpo),
        )
    )
    existente = result.scalars().first()
    if existente:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Já existe um grupo com esse nome")

    grupo = TagGroup(
        empresa_id=emp_uuid,
        nome=nome_limpo,
        cor=_normalizar_cor_tag(data.cor) if data.cor else None,
        ordem=data.ordem or 0,
    )
    db.add(grupo)

    try:
        await db.commit()
        await db.refresh(grupo)
        return {
            "id": str(grupo.id),
            "nome": grupo.nome,
            "cor": grupo.cor,
            "ordem": grupo.ordem,
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Erro ao criar grupo de tags: {str(e)}")


@router.put("/{empresa_id}/crm/tags/grupos/{grupo_id}", response_model=TagGroupResponse)
async def atualizar_tag_grupo(
    empresa_id: str,
    grupo_id: str,
    data: TagGroupUpdate,
    db: AsyncSession = Depends(get_db),
):
    try:
        emp_uuid = uuid.UUID(empresa_id)
        grupo_uuid = uuid.UUID(grupo_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID inválido")

    result = await db.execute(
        select(TagGroup).where(
            TagGroup.id == grupo_uuid,
            TagGroup.empresa_id == emp_uuid,
        )
    )
    grupo = result.scalars().first()
    if not grupo:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Grupo não encontrado")

    if data.nome is not None:
        nome_limpo = str(data.nome).strip()
        if not nome_limpo:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Nome do grupo é obrigatório")
        result_dup = await db.execute(
            select(TagGroup).where(
                TagGroup.empresa_id == emp_uuid,
                TagGroup.nome.ilike(nome_limpo),
                TagGroup.id != grupo.id,
            )
        )
        if result_dup.scalars().first():
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Já existe um grupo com esse nome")
        grupo.nome = nome_limpo
    if data.cor is not None:
        grupo.cor = _normalizar_cor_tag(data.cor) if data.cor else None
    if data.ordem is not None:
        grupo.ordem = data.ordem

    try:
        await db.commit()
        await db.refresh(grupo)
        return {
            "id": str(grupo.id),
            "nome": grupo.nome,
            "cor": grupo.cor,
            "ordem": grupo.ordem,
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Erro ao atualizar grupo de tags: {str(e)}")


@router.delete("/{empresa_id}/crm/tags/grupos/{grupo_id}", status_code=status.HTTP_204_NO_CONTENT)
async def excluir_tag_grupo(empresa_id: str, grupo_id: str, db: AsyncSession = Depends(get_db)):
    try:
        emp_uuid = uuid.UUID(empresa_id)
        grupo_uuid = uuid.UUID(grupo_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID inválido")

    result = await db.execute(
        select(TagGroup).where(
            TagGroup.id == grupo_uuid,
            TagGroup.empresa_id == emp_uuid,
        )
    )
    grupo = result.scalars().first()
    if not grupo:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Grupo não encontrado")

    try:
        await db.execute(
            update(TagCRM)
            .where(TagCRM.grupo_id == grupo.id, TagCRM.empresa_id == emp_uuid)
            .values(grupo_id=None)
        )
        await db.delete(grupo)
        await db.commit()
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Erro ao excluir grupo de tags: {str(e)}")


@router.get("/{empresa_id}/crm/tags/oficiais", response_model=List[TagCRMResponse])
async def listar_tags_crm_oficiais(empresa_id: str, db: AsyncSession = Depends(get_db)):
    try:
        emp_uuid = uuid.UUID(empresa_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID de empresa inválido")

    result = await db.execute(
        select(TagCRM)
        .where(TagCRM.empresa_id == emp_uuid)
        .order_by(TagCRM.nome.asc())
    )
    tags = result.scalars().all()
    return [
        {
            "id": str(tag.id),
            "nome": tag.nome,
            "cor": tag.cor or "#2563eb",
            "instrucao_ia": tag.instrucao_ia,
            "grupo_id": str(tag.grupo_id) if tag.grupo_id else None,
            "disparar_conversao_ads": bool(tag.disparar_conversao_ads),
            "criado_em": tag.criado_em.isoformat() if tag.criado_em else None,
        }
        for tag in tags
    ]


@router.post("/{empresa_id}/crm/tags/oficiais", response_model=TagCRMResponse, status_code=status.HTTP_201_CREATED)
async def criar_tag_crm_oficial(empresa_id: str, data: TagCRMCreate, db: AsyncSession = Depends(get_db)):
    try:
        emp_uuid = uuid.UUID(empresa_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID de empresa inválido")

    nome_limpo = str(data.nome or "").strip()
    if not nome_limpo:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Nome da tag é obrigatório")

    result = await db.execute(
        select(TagCRM).where(
            TagCRM.empresa_id == emp_uuid,
            TagCRM.nome.ilike(nome_limpo),
        )
    )
    existente = result.scalars().first()
    if existente:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Já existe uma tag oficial com esse nome")

    grupo_uuid = None
    if data.grupo_id:
        try:
            grupo_uuid = uuid.UUID(str(data.grupo_id))
        except ValueError:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="grupo_id inválido")
        result_grupo = await db.execute(
            select(TagGroup).where(
                TagGroup.id == grupo_uuid,
                TagGroup.empresa_id == emp_uuid,
            )
        )
        if not result_grupo.scalars().first():
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Grupo de tags não encontrado")

    tag = TagCRM(
        empresa_id=emp_uuid,
        grupo_id=grupo_uuid,
        nome=nome_limpo,
        cor=_normalizar_cor_tag(data.cor),
        instrucao_ia=(data.instrucao_ia or "").strip() or None,
        disparar_conversao_ads=bool(data.disparar_conversao_ads),
    )
    db.add(tag)

    try:
        await db.commit()
        await db.refresh(tag)
        return {
            "id": str(tag.id),
            "nome": tag.nome,
            "cor": tag.cor,
            "instrucao_ia": tag.instrucao_ia,
            "grupo_id": str(tag.grupo_id) if tag.grupo_id else None,
            "disparar_conversao_ads": bool(tag.disparar_conversao_ads),
            "criado_em": tag.criado_em.isoformat() if tag.criado_em else None,
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Erro ao criar tag oficial: {str(e)}")


@router.put("/{empresa_id}/crm/tags/oficiais/{tag_id}", response_model=TagCRMResponse)
async def atualizar_tag_crm_oficial(
    empresa_id: str,
    tag_id: str,
    data: TagCRMUpdate,
    db: AsyncSession = Depends(get_db),
):
    try:
        emp_uuid = uuid.UUID(empresa_id)
        tag_uuid = uuid.UUID(tag_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID inválido")

    result = await db.execute(
        select(TagCRM).where(
            TagCRM.id == tag_uuid,
            TagCRM.empresa_id == emp_uuid,
        )
    )
    tag = result.scalars().first()
    if not tag:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Tag oficial não encontrada")

    if data.nome is not None:
        nome_limpo = str(data.nome).strip()
        if not nome_limpo:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Nome da tag é obrigatório")
        result_dup = await db.execute(
            select(TagCRM).where(
                TagCRM.empresa_id == emp_uuid,
                TagCRM.nome.ilike(nome_limpo),
                TagCRM.id != tag.id,
            )
        )
        if result_dup.scalars().first():
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Já existe uma tag oficial com esse nome")
        tag.nome = nome_limpo
    if data.cor is not None:
        tag.cor = _normalizar_cor_tag(data.cor)
    if data.instrucao_ia is not None:
        tag.instrucao_ia = str(data.instrucao_ia).strip() or None
    if data.disparar_conversao_ads is not None:
        tag.disparar_conversao_ads = bool(data.disparar_conversao_ads)
    if data.grupo_id is not None:
        if data.grupo_id == "":
            tag.grupo_id = None
        else:
            try:
                grupo_uuid = uuid.UUID(str(data.grupo_id))
            except ValueError:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="grupo_id inválido")
            result_grupo = await db.execute(
                select(TagGroup).where(
                    TagGroup.id == grupo_uuid,
                    TagGroup.empresa_id == emp_uuid,
                )
            )
            grupo = result_grupo.scalars().first()
            if not grupo:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Grupo de tags não encontrado")
            tag.grupo_id = grupo.id

    try:
        await db.commit()
        await db.refresh(tag)
        return {
            "id": str(tag.id),
            "nome": tag.nome,
            "cor": tag.cor,
            "instrucao_ia": tag.instrucao_ia,
            "grupo_id": str(tag.grupo_id) if tag.grupo_id else None,
            "disparar_conversao_ads": bool(tag.disparar_conversao_ads),
            "criado_em": tag.criado_em.isoformat() if tag.criado_em else None,
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Erro ao atualizar tag oficial: {str(e)}")


@router.delete("/{empresa_id}/crm/tags/oficiais/{tag_id}", status_code=status.HTTP_204_NO_CONTENT)
async def excluir_tag_crm_oficial(empresa_id: str, tag_id: str, db: AsyncSession = Depends(get_db)):
    try:
        emp_uuid = uuid.UUID(empresa_id)
        tag_uuid = uuid.UUID(tag_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID inválido")

    result = await db.execute(
        select(TagCRM).where(
            TagCRM.id == tag_uuid,
            TagCRM.empresa_id == emp_uuid,
        )
    )
    tag = result.scalars().first()
    if not tag:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Tag oficial não encontrada")

    try:
        await db.delete(tag)
        await db.commit()
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Erro ao excluir tag oficial: {str(e)}")


@router.get("/{empresa_id}/crm/tags", response_model=List[str])
async def listar_tags_crm(empresa_id: str, db: AsyncSession = Depends(get_db)):
    try:
        return await listar_tags_oficiais_ou_existentes(empresa_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID de empresa inválido")


@router.get("/{empresa_id}/campanhas/templates", response_model=List[TemplateMensagemResponse])
async def listar_templates_mensagem(empresa_id: str, db: AsyncSession = Depends(get_db)):
    try:
        emp_uuid = uuid.UUID(empresa_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID de empresa inválido")

    result = await db.execute(
        select(TemplateMensagem)
        .where(TemplateMensagem.empresa_id == emp_uuid)
        .order_by(TemplateMensagem.criado_em.desc())
    )
    templates = result.scalars().all()
    return [
        {
            "id": str(template.id),
            "nome": template.nome,
            "texto_template": template.texto_template,
            "variaveis_esperadas": template.variaveis_esperadas or [],
            "criado_em": template.criado_em.isoformat() if template.criado_em else None,
        }
        for template in templates
    ]


@router.post("/{empresa_id}/campanhas/templates", response_model=TemplateMensagemResponse, status_code=status.HTTP_201_CREATED)
async def criar_template_mensagem(
    empresa_id: str,
    data: TemplateMensagemCreate,
    db: AsyncSession = Depends(get_db),
):
    try:
        emp_uuid = uuid.UUID(empresa_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID de empresa inválido")

    variaveis = data.variaveis_esperadas or extrair_variaveis_template(data.texto_template)
    template = TemplateMensagem(
        empresa_id=emp_uuid,
        nome=data.nome.strip(),
        texto_template=data.texto_template,
        variaveis_esperadas=[str(item).strip() for item in variaveis if str(item).strip()],
    )
    db.add(template)

    try:
        await db.commit()
        await db.refresh(template)
        return {
            "id": str(template.id),
            "nome": template.nome,
            "texto_template": template.texto_template,
            "variaveis_esperadas": template.variaveis_esperadas or [],
            "criado_em": template.criado_em.isoformat() if template.criado_em else None,
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Erro ao criar template: {str(e)}")


@router.put("/{empresa_id}/campanhas/templates/{template_id}", response_model=TemplateMensagemResponse)
async def atualizar_template_mensagem(
    empresa_id: str,
    template_id: str,
    data: TemplateMensagemUpdate,
    db: AsyncSession = Depends(get_db),
):
    try:
        emp_uuid = uuid.UUID(empresa_id)
        template_uuid = uuid.UUID(template_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID inválido")

    result = await db.execute(
        select(TemplateMensagem).where(
            TemplateMensagem.id == template_uuid,
            TemplateMensagem.empresa_id == emp_uuid,
        )
    )
    template = result.scalars().first()
    if not template:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Template não encontrado")

    texto_template_final = data.texto_template if data.texto_template is not None else template.texto_template
    if data.nome is not None:
        template.nome = data.nome.strip()
    if data.texto_template is not None:
        template.texto_template = data.texto_template
    if data.variaveis_esperadas is not None:
        template.variaveis_esperadas = [str(item).strip() for item in data.variaveis_esperadas if str(item).strip()]
    else:
        template.variaveis_esperadas = extrair_variaveis_template(texto_template_final)

    try:
        await db.commit()
        await db.refresh(template)
        return {
            "id": str(template.id),
            "nome": template.nome,
            "texto_template": template.texto_template,
            "variaveis_esperadas": template.variaveis_esperadas or [],
            "criado_em": template.criado_em.isoformat() if template.criado_em else None,
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Erro ao atualizar template: {str(e)}")


@router.delete("/{empresa_id}/campanhas/templates/{template_id}", status_code=status.HTTP_204_NO_CONTENT)
async def excluir_template_mensagem(
    empresa_id: str,
    template_id: str,
    db: AsyncSession = Depends(get_db),
):
    try:
        emp_uuid = uuid.UUID(empresa_id)
        template_uuid = uuid.UUID(template_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID inválido")

    result = await db.execute(
        select(TemplateMensagem).where(
            TemplateMensagem.id == template_uuid,
            TemplateMensagem.empresa_id == emp_uuid,
        )
    )
    template = result.scalars().first()
    if not template:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Template não encontrado")

    try:
        await db.delete(template)
        await db.commit()
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Erro ao excluir template: {str(e)}")


@router.get("/{empresa_id}/campanhas", response_model=List[CampanhaDisparoResponse])
async def listar_campanhas_disparo(empresa_id: str, db: AsyncSession = Depends(get_db)):
    try:
        emp_uuid = uuid.UUID(empresa_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID de empresa inválido")

    result = await db.execute(
        select(CampanhaDisparo)
        .where(CampanhaDisparo.empresa_id == emp_uuid)
        .options(selectinload(CampanhaDisparo.template))
        .order_by(CampanhaDisparo.criado_em.desc())
    )
    campanhas = result.scalars().all()
    return [
        {
            "id": str(campanha.id),
            "nome": campanha.nome,
            "template_id": str(campanha.template_id) if campanha.template_id else None,
            "template_nome": campanha.template.nome if campanha.template else None,
            "tags_alvo": campanha.tags_alvo or [],
            "data_agendamento": campanha.data_agendamento.isoformat() if campanha.data_agendamento else None,
            "status": campanha.status.value if hasattr(campanha.status, "value") else str(campanha.status),
            "criado_em": campanha.criado_em.isoformat() if campanha.criado_em else None,
        }
        for campanha in campanhas
    ]


@router.get("/{empresa_id}/campanhas/listas", response_model=List[CampanhaListaResponse])
async def listar_listas_campanhas(empresa_id: str, db: AsyncSession = Depends(get_db)):
    try:
        emp_uuid = uuid.UUID(empresa_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID de empresa inválido")

    result_tags = await db.execute(
        select(TagCRM)
        .where(TagCRM.empresa_id == emp_uuid)
        .order_by(TagCRM.nome.asc())
    )
    tags_oficiais = result_tags.scalars().all()

    result_leads = await db.execute(
        select(CRMLead.tags).where(CRMLead.empresa_id == emp_uuid)
    )
    leads_tags = result_leads.scalars().all()

    def _contar(nome_tag: str) -> int:
        nome_normalizado = str(nome_tag).strip().lower()
        total = 0
        for tags in leads_tags:
            tags_normalizadas = {str(tag).strip().lower() for tag in (tags or []) if str(tag).strip()}
            if nome_normalizado in tags_normalizadas:
                total += 1
        return total

    return [
        {
            "tag_id": str(tag.id),
            "tag": tag.nome,
            "cor": tag.cor or "#2563eb",
            "total_leads": _contar(tag.nome),
        }
        for tag in tags_oficiais
    ]


@router.get("/{empresa_id}/campanhas/listas/{tag_id}/leads", response_model=List[CampanhaListaLeadResponse])
async def listar_leads_da_lista_campanha(empresa_id: str, tag_id: str, db: AsyncSession = Depends(get_db)):
    try:
        emp_uuid = uuid.UUID(empresa_id)
        tag_uuid = uuid.UUID(tag_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID inválido")

    result_tag = await db.execute(
        select(TagCRM).where(
            TagCRM.id == tag_uuid,
            TagCRM.empresa_id == emp_uuid,
        )
    )
    tag = result_tag.scalars().first()
    if not tag:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Tag oficial não encontrada")

    result_leads = await db.execute(
        select(CRMLead)
        .where(CRMLead.empresa_id == emp_uuid)
        .options(selectinload(CRMLead.etapa))
        .order_by(CRMLead.criado_em.desc())
    )
    leads = result_leads.scalars().all()
    tag_normalizada = tag.nome.strip().lower()

    return [
        {
            "id": str(lead.id),
            "nome_contato": lead.nome_contato,
            "telefone_contato": lead.telefone_contato,
            "status": lead.etapa.nome if lead.etapa else None,
            "tags": lead.tags or [],
            "historico_resumo": lead.historico_resumo,
            "dados_adicionais": lead.dados_adicionais or {},
        }
        for lead in leads
        if tag_normalizada in {str(item).strip().lower() for item in (lead.tags or []) if str(item).strip()}
    ]


@router.get("/{empresa_id}/campanhas/preview", response_model=CampanhaPreviewResponse)
async def preview_campanha_disparo(
    empresa_id: str,
    template_id: str,
    tag: str,
    db: AsyncSession = Depends(get_db),
):
    try:
        preview = await gerar_preview_campanha(
            session=db,
            empresa_id=empresa_id,
            template_id=template_id,
            tag=tag,
        )
        return preview
    except ValueError as e:
        mensagem = str(e)
        status_code = status.HTTP_404_NOT_FOUND if "não encontrado" in mensagem.lower() else status.HTTP_400_BAD_REQUEST
        raise HTTPException(status_code=status_code, detail=mensagem)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erro ao gerar pré-visualização: {str(e)}",
        )


@router.post("/{empresa_id}/campanhas", response_model=CampanhaDisparoResponse, status_code=status.HTTP_201_CREATED)
async def criar_campanha_disparo(
    empresa_id: str,
    data: CampanhaDisparoCreate,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
):
    try:
        emp_uuid = uuid.UUID(empresa_id)
        template_uuid = uuid.UUID(data.template_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID inválido")

    result_empresa = await db.execute(select(Empresa).where(Empresa.id == emp_uuid))
    empresa = result_empresa.scalars().first()
    if not empresa:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Empresa não encontrada")
    if not empresa.conexao_disparo_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Configure a conexão padrão para disparos antes de criar uma campanha",
        )

    result_template = await db.execute(
        select(TemplateMensagem).where(
            TemplateMensagem.id == template_uuid,
            TemplateMensagem.empresa_id == emp_uuid,
        )
    )
    template = result_template.scalars().first()
    if not template:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Template não encontrado")

    tags_alvo = [str(tag).strip() for tag in (data.tags_alvo or []) if str(tag).strip()]
    if not tags_alvo:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Selecione ao menos uma tag alvo")

    result_leads = await db.execute(
        select(CRMLead.tags).where(CRMLead.empresa_id == emp_uuid)
    )
    total_leads = 0
    tags_alvo_normalizadas = {tag.lower() for tag in tags_alvo}
    for tags in result_leads.scalars().all():
        if not isinstance(tags, list):
            continue
        tags_lead = {str(tag).strip().lower() for tag in tags if str(tag).strip()}
        if tags_alvo_normalizadas.intersection(tags_lead):
            total_leads += 1

    if total_leads == 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Não há leads com a tag selecionada para este disparo",
        )

    campanha = CampanhaDisparo(
        empresa_id=emp_uuid,
        nome=data.nome.strip(),
        template_id=template.id,
        tags_alvo=tags_alvo,
        data_agendamento=datetime.fromisoformat(data.data_agendamento) if data.data_agendamento else datetime.utcnow(),
        status=CampanhaDisparoStatus.PENDENTE,
    )
    db.add(campanha)

    try:
        await db.commit()
        await db.refresh(campanha)
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Erro ao criar campanha: {str(e)}")

    background_tasks.add_task(processar_campanha_disparo, str(campanha.id))

    return {
        "id": str(campanha.id),
        "nome": campanha.nome,
        "template_id": str(campanha.template_id) if campanha.template_id else None,
        "template_nome": template.nome,
        "tags_alvo": campanha.tags_alvo or [],
        "data_agendamento": campanha.data_agendamento.isoformat() if campanha.data_agendamento else None,
        "status": campanha.status.value if hasattr(campanha.status, "value") else str(campanha.status),
        "criado_em": campanha.criado_em.isoformat() if campanha.criado_em else None,
    }


@router.put("/{empresa_id}/crm/etapas/{etapa_id}", status_code=status.HTTP_200_OK)
async def atualizar_etapa_crm(
    empresa_id: str,
    etapa_id: str,
    data: CRMEtapaUpdateRequest,
    db: AsyncSession = Depends(get_db)
):
    """
    Atualiza dados de uma etapa do CRM (nome, tipo e ordem).
    """
    try:
        emp_uuid = uuid.UUID(empresa_id)
        etapa_uuid = uuid.UUID(etapa_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID inválido")

    result = await db.execute(
        select(CRMEtapa)
        .join(CRMFunil, CRMEtapa.funil_id == CRMFunil.id)
        .where(CRMEtapa.id == etapa_uuid, CRMFunil.empresa_id == emp_uuid)
    )
    etapa = result.scalars().first()
    if not etapa:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Etapa não encontrada")

    if data.nome is not None:
        etapa.nome = data.nome
    if data.tipo is not None:
        etapa.tipo = data.tipo
    if data.ordem is not None:
        etapa.ordem = data.ordem

    try:
        await db.commit()
        await db.refresh(etapa)
        return {
            "id": str(etapa.id),
            "nome": etapa.nome,
            "tipo": etapa.tipo,
            "ordem": etapa.ordem
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Erro ao atualizar etapa: {str(e)}")


@router.put("/{empresa_id}/crm/leads/{lead_id}", status_code=status.HTTP_200_OK)
async def atualizar_lead_crm(
    empresa_id: str,
    lead_id: str,
    data: CRMLeadUpdateRequest,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db)
):
    """
    Atualiza dados de um lead do CRM, incluindo dados_adicionais.
    """
    if _lead_id_eh_simulador_ou_invalido(lead_id):
        return {
            "id": lead_id,
            "nome_contato": data.nome_contato or "[Simulador]",
            "telefone_contato": data.telefone_contato or SIMULADOR_LEAD_ID,
            "historico_resumo": data.historico_resumo or "",
            "etapa_id": data.etapa_id or None,
            "status_atendimento": data.status_atendimento or "aberto",
            "tags": data.tags or [],
            "dados_adicionais": data.dados_adicionais or {},
            "valor_conversao": data.valor_conversao or 0,
        }

    try:
        emp_uuid = uuid.UUID(empresa_id)
        lead_uuid = uuid.UUID(lead_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID inválido")

    result = await db.execute(select(CRMLead).where(CRMLead.id == lead_uuid, CRMLead.empresa_id == emp_uuid))
    lead = result.scalars().first()
    if not lead:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Lead não encontrado")

    if data.etapa_id is not None:
        if data.etapa_id == "":
            lead.etapa_id = None
        else:
            try:
                etapa_uuid = uuid.UUID(data.etapa_id)
            except ValueError:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="etapa_id inválido")

            result_etapa = await db.execute(
                select(CRMEtapa)
                .join(CRMFunil, CRMEtapa.funil_id == CRMFunil.id)
                .where(CRMEtapa.id == etapa_uuid, CRMFunil.empresa_id == emp_uuid)
            )
            etapa = result_etapa.scalars().first()
            if not etapa:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Etapa não encontrada para esta empresa")
            lead.etapa_id = etapa.id

    if data.nome_contato is not None:
        lead.nome_contato = data.nome_contato
    if data.telefone_contato is not None:
        lead.telefone_contato = data.telefone_contato
    if data.historico_resumo is not None:
        lead.historico_resumo = data.historico_resumo
    if data.status_atendimento is not None:
        status_limpo = str(data.status_atendimento or "").strip().lower()
        if status_limpo not in {"aberto", "concluido"}:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="status_atendimento inválido")
        lead.status_atendimento = status_limpo
    tags_para_notificar: list[TagCRM] = []
    if data.tags is not None:
        tags_atuais_norm = {str(tag).strip().lower() for tag in (lead.tags or []) if str(tag).strip()}
        tags_limpas = [str(tag).strip() for tag in data.tags if str(tag).strip()]
        lead.tags = tags_limpas
        tags_novas_aplicadas = [tag for tag in tags_limpas if tag.lower() not in tags_atuais_norm]
        tags_novas_norm = {tag.lower() for tag in tags_novas_aplicadas if tag}
        if tags_novas_norm:
            result_tags_disparo = await db.execute(
                select(TagCRM).where(
                    TagCRM.empresa_id == emp_uuid,
                    TagCRM.disparar_conversao_ads == True,
                )
            )
            tags_para_notificar = [
                tag
                for tag in result_tags_disparo.scalars().all()
                if str(tag.nome or "").strip().lower() in tags_novas_norm
            ]
    if data.dados_adicionais is not None:
        lead.dados_adicionais = data.dados_adicionais
    if data.valor_conversao is not None:
        lead.valor_conversao = float(data.valor_conversao)

    try:
        await db.commit()
        for tag in tags_para_notificar:
            background_tasks.add_task(notificar_conversao_ads, str(lead.id), str(tag.nome), db)
        await db.refresh(lead)
        return {
            "id": str(lead.id),
            "nome_contato": lead.nome_contato,
            "telefone_contato": lead.telefone_contato,
            "historico_resumo": lead.historico_resumo,
            "etapa_id": str(lead.etapa_id) if lead.etapa_id else None,
            "status_atendimento": lead.status_atendimento,
            "tags": lead.tags or [],
            "dados_adicionais": lead.dados_adicionais or {},
            "valor_conversao": lead.valor_conversao,
        }
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Erro ao atualizar lead: {str(e)}")


@router.post("/{empresa_id}/leads/{lead_id}/transferir_manual", status_code=status.HTTP_200_OK)
async def transferir_lead_manual(
    empresa_id: str,
    lead_id: str,
    data: TransferenciaManualRequest,
):
    if _lead_id_eh_simulador_ou_invalido(lead_id):
        return {"success": True, "detail": "Ação simulada"}

    resultado = await executar_transferencia_atendimento(
        empresa_id=empresa_id,
        lead_id=lead_id,
        destino_id=data.destino_id,
        resumo_conversa="Transferência manual realizada pelo atendente no painel",
    )

    if resultado.lower().startswith("erro ao transferir atendimento"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=resultado)

    return {"success": True, "detail": resultado}


@router.post("/{empresa_id}/leads/importar", status_code=status.HTTP_201_CREATED)
async def importar_leads(
    empresa_id: str,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    tags_iniciais: str = Form("[]"),
    db: AsyncSession = Depends(get_db),
):
    try:
        emp_uuid = uuid.UUID(empresa_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID de empresa inválido")

    if not file.filename:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Arquivo inválido")

    try:
        tags_iniciais_lista = normalizar_tags(json.loads(tags_iniciais or "[]"))
        file_bytes = await file.read()
        headers, rows = _load_spreadsheet_rows(file_bytes, file.filename)
        nome_col, telefone_col = _detect_lead_columns(headers)
        if not nome_col or not telefone_col:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Não encontrei as colunas de Nome e Telefone no arquivo enviado.",
            )

        etapa_inicial_id = await _garantir_etapa_inicial_crm(db, emp_uuid)
        inseridos = 0
        atualizados = 0
        notificacoes_ads_pendentes: list[tuple[str, str]] = []

        result_tags_disparo = await db.execute(
            select(TagCRM).where(
                TagCRM.empresa_id == emp_uuid,
                TagCRM.disparar_conversao_ads == True,
            )
        )
        tags_disparo_map = {
            str(tag.nome or "").strip().lower(): tag
            for tag in result_tags_disparo.scalars().all()
            if str(tag.nome or "").strip()
        }

        for row in rows:
            nome = str((row or {}).get(nome_col, "")).strip()
            telefone = str((row or {}).get(telefone_col, "")).strip()
            if not nome and not telefone:
                continue

            lead = None
            if telefone:
                result_lead = await db.execute(
                    select(CRMLead).where(
                        CRMLead.empresa_id == emp_uuid,
                        CRMLead.telefone_contato == telefone,
                    )
                )
                lead = result_lead.scalars().first()

            if lead:
                if nome:
                    lead.nome_contato = nome
                tags_antes = {str(tag).strip().lower() for tag in (lead.tags or []) if str(tag).strip()}
                lead.tags = normalizar_tags((lead.tags or []) + tags_iniciais_lista)
                tags_novas = [tag for tag in (lead.tags or []) if str(tag).strip().lower() not in tags_antes]
                for tag_nome in tags_novas:
                    tag_obj = tags_disparo_map.get(str(tag_nome).strip().lower())
                    if tag_obj:
                        notificacoes_ads_pendentes.append((str(lead.id), str(tag_obj.nome)))
                atualizados += 1
            else:
                novo_lead = CRMLead(
                    empresa_id=emp_uuid,
                    etapa_id=etapa_inicial_id,
                    nome_contato=nome or telefone or "Lead importado",
                    telefone_contato=telefone or None,
                    tags=tags_iniciais_lista,
                )
                db.add(novo_lead)
                await db.flush()
                for tag_nome in tags_iniciais_lista:
                    tag_obj = tags_disparo_map.get(str(tag_nome).strip().lower())
                    if tag_obj:
                        notificacoes_ads_pendentes.append((str(novo_lead.id), str(tag_obj.nome)))
                inseridos += 1

        await db.commit()
        for lead_id_notificacao, tag_nome_notificacao in notificacoes_ads_pendentes:
            background_tasks.add_task(notificar_conversao_ads, lead_id_notificacao, tag_nome_notificacao, db)
        return {
            "status": "sucesso",
            "mensagem": "Importação concluída.",
            "colunas_detectadas": headers,
            "tags_iniciais": tags_iniciais_lista,
            "inseridos": inseridos,
            "atualizados": atualizados,
            "total_processado": inseridos + atualizados,
        }
    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Erro ao importar leads: {str(e)}")


@router.get("/{empresa_id}/exportar-leads")
async def exportar_leads_csv(empresa_id: str, db: AsyncSession = Depends(get_db)):
    """
    Exporta os leads em CSV (UTF-8-SIG), expandindo dados_adicionais em colunas extras.
    """
    try:
        emp_uuid = uuid.UUID(empresa_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID de empresa inválido")

    result = await db.execute(
        select(CRMLead)
        .where(CRMLead.empresa_id == emp_uuid)
        .options(selectinload(CRMLead.etapa))
        .order_by(CRMLead.criado_em.desc())
    )
    leads = result.scalars().all()

    extra_keys: set[str] = set()
    for lead in leads:
        if isinstance(lead.dados_adicionais, dict):
            extra_keys.update(str(k) for k in lead.dados_adicionais.keys())

    base_columns = [
        "id",
        "nome_contato",
        "telefone_contato",
        "historico_resumo",
        "etapa_id",
        "etapa_nome",
        "etapa_tipo",
        "tags",
        "criado_em",
    ]
    fieldnames = base_columns + sorted(extra_keys)

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()

    for lead in leads:
        row = {
            "id": str(lead.id),
            "nome_contato": lead.nome_contato,
            "telefone_contato": lead.telefone_contato or "",
            "historico_resumo": lead.historico_resumo or "",
            "etapa_id": str(lead.etapa_id) if lead.etapa_id else "",
            "etapa_nome": lead.etapa.nome if lead.etapa else "",
            "etapa_tipo": lead.etapa.tipo if lead.etapa else "",
            "tags": ", ".join(lead.tags or []),
            "criado_em": lead.criado_em.isoformat() if lead.criado_em else "",
        }
        if isinstance(lead.dados_adicionais, dict):
            for key in extra_keys:
                value = lead.dados_adicionais.get(key, "")
                row[key] = "" if value is None else str(value)
        writer.writerow(row)

    csv_bytes = output.getvalue().encode("utf-8-sig")
    output.close()

    return Response(
        content=csv_bytes,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=leads_{empresa_id}.csv"},
    )


DIAS_SEMANA_ORDENADOS = ("seg", "ter", "qua", "qui", "sex", "sab", "dom")


class AgendaConfigUpdateRequest(BaseModel):
    agenda_config: Dict[str, Any]


def _parse_horario_hhmm(valor: str, campo: str) -> datetime:
    try:
        return datetime.strptime(str(valor).strip(), "%H:%M")
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Campo '{campo}' deve estar no formato HH:MM."
        ) from exc


def _dias_funcionamento_base() -> Dict[str, Dict[str, Any]]:
    return {
        dia: {"aberto": False, "inicio": None, "fim": None}
        for dia in DIAS_SEMANA_ORDENADOS
    }


def _normalizar_dias_funcionamento(
    dias_brutos: Any,
    horario_inicio_legacy: Any = None,
    horario_fim_legacy: Any = None,
) -> Dict[str, Dict[str, Any]]:
    dias_normalizados = _dias_funcionamento_base()

    # Formato novo: {"seg": {"aberto": true, "inicio": "08:00", "fim": "18:00"}, ...}
    if isinstance(dias_brutos, dict) and any(chave in dias_brutos for chave in DIAS_SEMANA_ORDENADOS):
        for dia in DIAS_SEMANA_ORDENADOS:
            item = dias_brutos.get(dia) or {}
            if not isinstance(item, dict):
                continue
            aberto = bool(item.get("aberto", False))
            if not aberto:
                dias_normalizados[dia] = {"aberto": False, "inicio": None, "fim": None}
                continue
            inicio = item.get("inicio")
            fim = item.get("fim")
            if isinstance(inicio, str) and isinstance(fim, str):
                dias_normalizados[dia] = {"aberto": True, "inicio": inicio.strip(), "fim": fim.strip()}
        return dias_normalizados

    # Formato legado: {"dias": ["seg", "ter", ...]}
    dias_legado = dias_brutos.get("dias", []) if isinstance(dias_brutos, dict) else []
    inicio_legado = horario_inicio_legacy.strftime("%H:%M") if horario_inicio_legacy else "08:00"
    fim_legado = horario_fim_legacy.strftime("%H:%M") if horario_fim_legacy else "18:00"
    for dia in dias_legado:
        dia_norm = str(dia).strip().lower()
        if dia_norm in dias_normalizados:
            dias_normalizados[dia_norm] = {
                "aberto": True,
                "inicio": inicio_legado,
                "fim": fim_legado,
            }
    return dias_normalizados


def _validar_e_normalizar_dias_payload(dias_payload: Any) -> Dict[str, Dict[str, Any]]:
    if not isinstance(dias_payload, dict):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="agenda_config.dias_funcionamento deve ser um objeto com os dias da semana."
        )

    chaves_recebidas = set(dias_payload.keys())
    chaves_esperadas = set(DIAS_SEMANA_ORDENADOS)
    if chaves_recebidas != chaves_esperadas:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="agenda_config.dias_funcionamento deve conter exatamente as chaves: seg, ter, qua, qui, sex, sab, dom."
        )

    resultado: Dict[str, Dict[str, Any]] = {}
    for dia in DIAS_SEMANA_ORDENADOS:
        item = dias_payload.get(dia)
        if not isinstance(item, dict):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"agenda_config.dias_funcionamento.{dia} deve ser um objeto."
            )

        aberto = item.get("aberto", None)
        if not isinstance(aberto, bool):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"agenda_config.dias_funcionamento.{dia}.aberto deve ser boolean."
            )

        inicio = item.get("inicio")
        fim = item.get("fim")

        if aberto:
            if not isinstance(inicio, str) or not isinstance(fim, str):
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"agenda_config.dias_funcionamento.{dia}.inicio e agenda_config.dias_funcionamento.{dia}.fim devem ser strings HH:MM quando aberto=true."
                )
            dt_inicio = _parse_horario_hhmm(inicio, f"agenda_config.dias_funcionamento.{dia}.inicio")
            dt_fim = _parse_horario_hhmm(fim, f"agenda_config.dias_funcionamento.{dia}.fim")
            if dt_inicio >= dt_fim:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"agenda_config.dias_funcionamento.{dia}.inicio deve ser menor que agenda_config.dias_funcionamento.{dia}.fim."
                )
            resultado[dia] = {"aberto": True, "inicio": dt_inicio.strftime("%H:%M"), "fim": dt_fim.strftime("%H:%M")}
            continue

        # Quando fechado, o padrão aceito é null/null
        if inicio is not None or fim is not None:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"agenda_config.dias_funcionamento.{dia}.inicio e agenda_config.dias_funcionamento.{dia}.fim devem ser null quando aberto=false."
            )
        resultado[dia] = {"aberto": False, "inicio": None, "fim": None}

    return resultado


def _validar_e_normalizar_excecoes_payload(excecoes_payload: Any) -> List[Dict[str, Any]]:
    if excecoes_payload is None:
        return []
    if not isinstance(excecoes_payload, list):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="agenda_config.excecoes deve ser uma lista."
        )

    excecoes_normalizadas: List[Dict[str, Any]] = []
    for idx, item in enumerate(excecoes_payload):
        if not isinstance(item, dict):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"agenda_config.excecoes[{idx}] deve ser um objeto."
            )

        data_str = str(item.get("data") or "").strip()
        titulo_str = str(item.get("titulo") or "").strip()
        aberto = item.get("aberto")

        if not data_str:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"agenda_config.excecoes[{idx}].data é obrigatório."
            )
        try:
            datetime.strptime(data_str, "%Y-%m-%d")
        except ValueError as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"agenda_config.excecoes[{idx}].data deve estar no formato YYYY-MM-DD."
            ) from exc

        if not titulo_str:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"agenda_config.excecoes[{idx}].titulo é obrigatório."
            )
        if not isinstance(aberto, bool):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"agenda_config.excecoes[{idx}].aberto deve ser boolean."
            )

        inicio = item.get("inicio")
        fim = item.get("fim")
        if aberto:
            if not isinstance(inicio, str) or not isinstance(fim, str):
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"agenda_config.excecoes[{idx}].inicio e fim devem ser strings HH:MM quando aberto=true."
                )
            dt_inicio = _parse_horario_hhmm(inicio, f"agenda_config.excecoes[{idx}].inicio")
            dt_fim = _parse_horario_hhmm(fim, f"agenda_config.excecoes[{idx}].fim")
            if dt_inicio >= dt_fim:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"agenda_config.excecoes[{idx}].inicio deve ser menor que fim."
                )
            excecoes_normalizadas.append(
                {
                    "data": data_str,
                    "titulo": titulo_str,
                    "aberto": True,
                    "inicio": dt_inicio.strftime("%H:%M"),
                    "fim": dt_fim.strftime("%H:%M"),
                }
            )
        else:
            if inicio is not None or fim is not None:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"agenda_config.excecoes[{idx}].inicio e fim devem ser null quando aberto=false."
                )
            excecoes_normalizadas.append(
                {
                    "data": data_str,
                    "titulo": titulo_str,
                    "aberto": False,
                    "inicio": None,
                    "fim": None,
                }
            )

    return excecoes_normalizadas


def _serialize_agenda_config(config: AgendaConfiguracao | None) -> dict | None:
    if not config:
        return None
    dias_obj = _normalizar_dias_funcionamento(
        config.dias_funcionamento,
        horario_inicio_legacy=config.horario_inicio,
        horario_fim_legacy=config.horario_fim,
    )
    return {
        "dias_funcionamento": dias_obj,
        "excecoes": config.excecoes if isinstance(config.excecoes, list) else [],
        "duracao_minutos": config.duracao_slot_minutos
    }


@router.put("/{empresa_id}/agenda", status_code=status.HTTP_200_OK)
async def atualizar_agenda(
    empresa_id: str,
    payload: AgendaConfigUpdateRequest,
    db: AsyncSession = Depends(get_db),
):
    """
    Upsert da configuração da agenda por dia da semana.
    A fonte de verdade é o JSONB dias_funcionamento.
    """
    try:
        try:
            emp_uuid = uuid.UUID(empresa_id)
        except ValueError as exc:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID de empresa inválido") from exc

        agenda_config = payload.agenda_config if isinstance(payload.agenda_config, dict) else {}
        # Compatibilidade: aceita tanto agenda_config direto com seg..dom quanto
        # agenda_config.dias_funcionamento com o mesmo conteúdo.
        dias_payload = (
            agenda_config.get("dias_funcionamento")
            if "dias_funcionamento" in agenda_config
            else agenda_config
        )
        dias_normalizados = _validar_e_normalizar_dias_payload(dias_payload)
        excecoes_normalizadas = _validar_e_normalizar_excecoes_payload(agenda_config.get("excecoes"))

        result_config = await db.execute(
            select(AgendaConfiguracao).where(AgendaConfiguracao.empresa_id == emp_uuid)
        )
        config = result_config.scalars().first()

        if config:
            config.dias_funcionamento = dias_normalizados
            config.excecoes = excecoes_normalizadas
            config.horario_inicio = None
            config.horario_fim = None
        else:
            config = AgendaConfiguracao(
                empresa_id=emp_uuid,
                dias_funcionamento=dias_normalizados,
                excecoes=excecoes_normalizadas,
                horario_inicio=None,
                horario_fim=None,
            )
            db.add(config)

        await db.commit()
        await db.refresh(config)

        return _serialize_agenda_config(config)
    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erro ao atualizar configuração da Agenda: {str(e)}"
        )


@router.get("/{empresa_id}/agenda")
async def obter_agenda(empresa_id: str, db: AsyncSession = Depends(get_db)):
    """
    Retorna as configurações da agenda e a lista de agendamentos futuros da empresa.
    """
    try:
        # Busca config da agenda
        result_config = await db.execute(
            select(AgendaConfiguracao).where(AgendaConfiguracao.empresa_id == empresa_id)
        )
        config = result_config.scalars().first()

        # Busca os agendamentos futuros (onde data_hora_inicio >= now, mas aqui para mockup pegaremos todos)
        # e dá load do Lead associado para mostrar o nome
        result_agendamentos = await db.execute(
            select(AgendamentoLocal)
            .where(AgendamentoLocal.empresa_id == empresa_id)
            .options(selectinload(AgendamentoLocal.lead))
            .order_by(AgendamentoLocal.data_hora_inicio)
        )
        agendamentos_db = result_agendamentos.scalars().all()

        resposta_config = _serialize_agenda_config(config)

        lista_agendamentos = []
        for ag in agendamentos_db:
            lista_agendamentos.append({
                "id": str(ag.id),
                "data_hora_inicio": ag.data_hora_inicio.isoformat() if ag.data_hora_inicio else None,
                "data_hora_fim": ag.data_hora_fim.isoformat() if ag.data_hora_fim else None,
                "status": ag.status,
                "lead": {
                    "nome_contato": ag.lead.nome_contato if ag.lead else "Desconhecido"
                }
            })

        return {
            "config": resposta_config,
            "agendamentos": lista_agendamentos
        }

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erro ao carregar dados da Agenda: {str(e)}"
        )


@router.put("/{empresa_id}/credenciais")
async def atualizar_credenciais(empresa_id: str, credenciais: EvolutionCredentials, db: AsyncSession = Depends(get_db)):
    """
    Atualiza as credenciais da Evolution API para uma determinada empresa (Super Admin).
    """
    try:
        result = await db.execute(select(Empresa).where(Empresa.id == empresa_id))
        empresa = result.scalars().first()
        
        if not empresa:
            raise HTTPException(status_code=404, detail="Empresa não encontrada")
            
        novas_credenciais = {
            "evolution_url": credenciais.evolution_url,
            "evolution_apikey": credenciais.evolution_apikey,
            "evolution_instance": credenciais.evolution_instance,
            "openai_api_key": credenciais.openai_api_key
        }
        
        # Merge de credenciais caso haja (no momento subscreve tudo do WhatsApp channel)
        # O model JSON admite dicionário
        empresa.credenciais_canais = novas_credenciais
        
        await db.commit()
        await db.refresh(empresa)
        
        return {"mensagem": "Credenciais atualizadas com sucesso", "credenciais": empresa.credenciais_canais}
        
    except HTTPException as e:
        raise e
    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erro ao atualizar credenciais: {str(e)}"
        )


class SimuladorRequest(BaseModel):
    mensagem: str
    sessao_id: str

@router.post("/{empresa_id}/simulador", status_code=status.HTTP_202_ACCEPTED)
async def simulador_chat(
    empresa_id: str,
    payload: SimuladorRequest,
    background_tasks: BackgroundTasks
):
    """
    Fire-and-Forget: salva a mensagem no Redis via handle_debouncer em background
    e retorna 202 imediatamente. A resposta da IA fica disponível via GET /simulador/resposta/{sessao_id}.
    """
    from app.api.utils import handle_debouncer
    from app.api.main import redis_client

    # Limpa qualquer resposta anterior desta sessão antes de iniciar novo ciclo
    await redis_client.delete(f"sim_resp:{payload.sessao_id}")

    msg = StandardMessage(
        empresa_id=empresa_id,
        canal="simulador",
        identificador_origem=payload.sessao_id,
        texto_mensagem=payload.mensagem,
        is_human_agent=False
    )
    background_tasks.add_task(handle_debouncer, msg)
    return {"status": "received", "sessao_id": payload.sessao_id}


@router.get("/{empresa_id}/simulador/resposta/{sessao_id}")
async def simulador_poll_resposta(empresa_id: str, sessao_id: str):
    """
    Long-polling: o frontend chama esta rota a cada 2s até receber status='concluido'.
    Quando o LangGraph finaliza, ele grava a resposta no Redis com chave sim_resp:{sessao_id}.
    """
    from app.api.main import redis_client

    resposta_raw = await redis_client.get(f"sim_resp:{sessao_id}")
    if resposta_raw is None:
        return {"status": "processando", "resposta": None}

    resposta = resposta_raw if isinstance(resposta_raw, str) else resposta_raw.decode("utf-8")
    # Consome a entrada após leitura para não poluir Redis
    await redis_client.delete(f"sim_resp:{sessao_id}")
    return {"status": "concluido", "resposta": resposta}

@router.delete("/{empresa_id}/simulador/reset", status_code=status.HTTP_204_NO_CONTENT)
async def resetar_simulador(empresa_id: str, db: AsyncSession = Depends(get_db)):
    """
    Remove fisicamente o Lead Mock do simulador e todo o seu histórico.
    """
    import uuid
    from db.models import CRMLead
    try:
        emp_uuid = uuid.UUID(empresa_id)
    except ValueError:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID de empresa inválido")

    await _limpar_estado_simulador_redis()

    result = await db.execute(select(CRMLead).where(
        CRMLead.empresa_id == emp_uuid,
        CRMLead.telefone_contato == "ID_TESTE_SIMULADOR"
    ))
    lead = result.scalars().first()
    
    if not lead:
        return Response(status_code=status.HTTP_204_NO_CONTENT)
        
    try:
        await db.delete(lead)
        await db.commit()
        return Response(status_code=status.HTTP_204_NO_CONTENT)
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Erro ao resetar simulador: {str(e)}")

@router.delete("/{empresa_id}/leads/{lead_id}", status_code=status.HTTP_204_NO_CONTENT)
async def deletar_lead(empresa_id: str, lead_id: str, db: AsyncSession = Depends(get_db)):
    """
    Remove fisicamente um Lead e todo o seu histórico de mensagens do banco de dados (Hard Delete).
    """
    if _lead_id_eh_simulador_ou_invalido(lead_id):
        await _limpar_estado_simulador_redis(lead_id if lead_id == SIMULADOR_LEAD_ID else SIMULADOR_LEAD_ID)
        return Response(status_code=status.HTTP_204_NO_CONTENT)

    emp_uuid = _parse_uuid_or_none(empresa_id)
    if not emp_uuid:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="ID de empresa inválido")

    l_uuid = _parse_uuid_or_none(lead_id)

    result = await db.execute(select(CRMLead).where(CRMLead.id == l_uuid, CRMLead.empresa_id == emp_uuid))
    lead = result.scalars().first()
    
    if not lead:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Lead não encontrado ou não pertence a esta empresa")
        
    try:
        # Limpeza explícita de dependências para evitar falha por FK em bases legadas sem ON DELETE efetivo.
        await db.execute(
            delete(MensagemHistorico).where(MensagemHistorico.lead_id == lead.id)
        )
        await db.execute(
            delete(HistoricoTransferencia).where(HistoricoTransferencia.lead_id == lead.id)
        )
        await db.execute(
            update(AgendamentoLocal)
            .where(AgendamentoLocal.lead_id == lead.id)
            .values(lead_id=None)
        )

        await db.delete(lead)
        await db.commit()
        return Response(status_code=status.HTTP_204_NO_CONTENT)
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"Erro ao deletar lead: {str(e)}")

